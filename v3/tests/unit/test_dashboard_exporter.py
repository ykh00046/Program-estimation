"""DashboardExporter.export_excel 단위 테스트 (PDCA #25). win32com 무의존."""
import importlib
import os
import sys
import tempfile
import unittest
from datetime import datetime
from unittest.mock import MagicMock


def _ensure_dependencies() -> None:
    try:
        importlib.import_module("openpyxl")
    except ModuleNotFoundError:
        raise unittest.SkipTest("DashboardExporter tests require openpyxl")


_ensure_dependencies()

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from openpyxl import load_workbook

from models.dashboard_exporter import DashboardExporter


def _make_dm(monthly=None, top=None, workers=None, recipes=None):
    dm = MagicMock()
    dm.get_monthly_production_stats.return_value = monthly if monthly is not None else []
    dm.get_top_materials.return_value = top if top is not None else []
    dm.get_worker_stats.return_value = workers if workers is not None else []
    dm.get_recipe_frequency.return_value = recipes if recipes is not None else []
    return dm


def _all_cell_text(path):
    wb = load_workbook(path)
    ws = wb.active
    texts = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is not None:
                texts.append(str(v))
    return texts


class TestDashboardExporter(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.out = self._tmp.name

    def tearDown(self):
        self._tmp.cleanup()

    def test_export_excel_creates_file(self):
        exp = DashboardExporter(_make_dm(), output_folder=self.out)
        path = exp.export_excel("2026-05-01", "2026-05-31")
        self.assertIsNotNone(path)
        self.assertTrue(os.path.exists(path))
        load_workbook(path)  # 재오픈 가능

    def test_sections_and_headers(self):
        exp = DashboardExporter(_make_dm(), output_folder=self.out)
        path = exp.export_excel(None, None)
        texts = _all_cell_text(path)
        for marker in ("[요약]", "[월별 생산량 (최근 6개월)]", "[자재 사용량 TOP 10]", "[작업자 통계]"):
            self.assertIn(marker, texts)
        for header in ("연월", "품목코드", "작업자", "총사용량(g)"):
            self.assertIn(header, texts)

    def test_top_materials_rows(self):
        top = [
            {"material_code": "C1", "material_name": "M1", "total_actual": 50.0, "use_count": 2},
            {"material_code": "C2", "material_name": "M2", "total_actual": 30.0, "use_count": 1},
        ]
        exp = DashboardExporter(_make_dm(top=top), output_folder=self.out)
        texts = _all_cell_text(exp.export_excel("2026-05-01", "2026-05-31"))
        self.assertIn("C1", texts)
        self.assertIn("M2", texts)

    def test_empty_data_safe(self):
        exp = DashboardExporter(_make_dm(), output_folder=self.out)
        path = exp.export_excel(None, None)
        self.assertIsNotNone(path)
        self.assertIn("기록 없음", _all_cell_text(path))

    def test_kpi_workers_and_recipes_counts(self):
        dm = _make_dm(workers=[{"worker": "a"}, {"worker": "b"}],
                      recipes=[{"recipe_name": "r1"}, {"recipe_name": "r2"}, {"recipe_name": "r3"}])
        exp = DashboardExporter(dm, output_folder=self.out)
        texts = _all_cell_text(exp.export_excel(None, None))
        # KPI 요약에 활성 작업자 2 / 누적 레시피 3 기록
        self.assertIn("활성 작업자 수", texts)
        self.assertIn("2", texts)
        self.assertIn("3", texts)

    def test_period_label(self):
        self.assertEqual(DashboardExporter._period_label(None, None), "전체")
        self.assertEqual(DashboardExporter._period_label("2026-05-01", "2026-05-31"), "2026-05-01~2026-05-31")

    def test_filename_uses_period(self):
        exp = DashboardExporter(_make_dm(), output_folder=self.out)
        path = exp.export_excel("2026-05-01", "2026-05-31")
        self.assertTrue(os.path.basename(path).startswith("대시보드_2026-05-01_2026-05-31"))


if __name__ == "__main__":
    unittest.main()
