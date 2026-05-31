"""대시보드 보고서 Excel/PDF 출력 (PDCA #25).

#17 대시보드의 집계(KPI/월별/자재TOP/작업자)를 단일 시트 워크북으로 내보낸다.
export_excel은 openpyxl만 의존(win32com/fitz 무관)하여 단독 테스트가 가능하고,
export_pdf는 그 Excel을 win32com으로 일반 PDF(스캔효과 없음)로 변환한다.
"""
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, Side

from config.config_manager import config
from utils.logger import logger

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


class DashboardExporter:
    """대시보드 집계를 Excel/PDF 보고서로 출력."""

    def __init__(self, data_manager, output_folder: Optional[str] = None) -> None:
        self.data_manager = data_manager
        if output_folder is None:
            base = config.get("paths.output", "실적서")
            output_folder = os.path.join(base, "dashboard")
        self.output_folder = output_folder

    # ------------------------------------------------------------------
    # Public
    # ------------------------------------------------------------------

    def export_excel(self, start_date: Optional[str], end_date: Optional[str],
                     filename: Optional[str] = None) -> Optional[str]:
        """4섹션 단일 시트 워크북을 생성하고 경로를 반환(실패 시 None)."""
        try:
            os.makedirs(self.output_folder, exist_ok=True)
            wb = self._build_workbook(start_date, end_date)
            if filename is None:
                label = self._period_label(start_date, end_date).replace("~", "_")
                filename = f"대시보드_{label}.xlsx"
            path = os.path.join(self.output_folder, filename)
            wb.save(path)
            logger.info(f"대시보드 Excel 생성: {path}")
            return path
        except (OSError, ValueError) as e:
            logger.error(f"대시보드 Excel 출력 오류: {e}", exc_info=True)
            return None

    def export_pdf(self, start_date: Optional[str], end_date: Optional[str]) -> Optional[str]:
        """export_excel 후 win32com으로 일반 PDF로 변환(실패 시 None)."""
        xlsx = self.export_excel(start_date, end_date)
        if not xlsx:
            return None
        pdf = xlsx[:-5] + ".pdf"
        try:
            self._excel_to_pdf(xlsx, pdf)
            logger.info(f"대시보드 PDF 생성: {pdf}")
            return pdf
        except Exception as e:  # noqa: BLE001 — win32com COM 오류 광역 흡수
            logger.error(f"대시보드 PDF 변환 오류: {e}", exc_info=True)
            return None

    # ------------------------------------------------------------------
    # Internal
    # ------------------------------------------------------------------

    def _build_workbook(self, start_date: Optional[str], end_date: Optional[str]) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "대시보드"

        title_cell = ws.cell(row=1, column=1, value="배합 이력 대시보드 보고서")
        title_cell.font = Font(bold=True, size=14)
        ws.cell(
            row=2, column=1,
            value=f"기간: {self._period_label(start_date, end_date)} | "
                  f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )

        kpi = self._compute_kpis()
        row = 4
        row = self._write_section(
            ws, row, "[요약]", None,
            [
                ("당월 생산 건수", kpi["records"]),
                ("당월 총 배합량(g)", kpi["amount"]),
                ("활성 작업자 수", kpi["workers"]),
                ("누적 레시피 종류", kpi["recipes"]),
            ],
        )

        monthly = self.data_manager.get_monthly_production_stats(months=6)
        monthly_title_row = row
        row = self._write_section(
            ws, row, "[월별 생산량 (최근 6개월)]",
            ["연월", "생산건수", "총배합량(g)"],
            [
                (r.get("year_month") or "", int(r.get("record_count") or 0), float(r.get("total_amount") or 0))
                for r in monthly
            ],
        )
        if monthly:
            self._add_monthly_chart(ws, monthly_title_row, len(monthly))

        top = self.data_manager.get_top_materials(limit=10, start_date=start_date, end_date=end_date)
        row = self._write_section(
            ws, row, "[자재 사용량 TOP 10]",
            ["순위", "품목코드", "품목명", "총사용량(g)", "사용횟수"],
            [
                (i + 1, r.get("material_code") or "", r.get("material_name") or "",
                 float(r.get("total_actual") or 0), int(r.get("use_count") or 0))
                for i, r in enumerate(top)
            ],
        )

        workers = self.data_manager.get_worker_stats(start_date=start_date, end_date=end_date)
        row = self._write_section(
            ws, row, "[작업자 통계]",
            ["작업자", "건수", "총량(g)", "평균(g)"],
            [
                (r.get("worker") or "", int(r.get("record_count") or 0),
                 float(r.get("total_amount") or 0), float(r.get("avg_amount") or 0))
                for r in workers
            ],
        )
        return wb

    def _write_section(self, ws: Any, row: int, title: str,
                       headers: Optional[List[str]], rows: List[tuple]) -> int:
        """섹션 제목 + (선택)헤더 + 데이터행을 기록하고 다음 시작행을 반환."""
        ws.cell(row=row, column=1, value=title).font = Font(bold=True)
        row += 1
        if headers:
            for col, head in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=head)
                cell.font = Font(bold=True)
                cell.border = _BORDER
                cell.alignment = Alignment(horizontal="center")
            row += 1
        if not rows:
            ws.cell(row=row, column=1, value="기록 없음")
            return row + 2
        for data in rows:
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                if headers:
                    cell.border = _BORDER
            row += 1
        return row + 1

    def _add_monthly_chart(self, ws: Any, title_row: int, count: int) -> None:
        """월별 섹션 데이터(C열=총배합량, A열=연월)를 참조하는 막대 차트를 추가한다.

        title_row 기준: header_row=+1, data=+2..+1+count (_write_section 레이아웃).
        """
        header_row = title_row + 1
        data_start = title_row + 2
        data_end = data_start + count - 1

        chart = BarChart()
        chart.type = "col"
        chart.title = "월별 총 배합량(g)"
        chart.legend = None
        chart.height = 7
        chart.width = 14

        data = Reference(ws, min_col=3, min_row=header_row, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"F{title_row}")

    def _compute_kpis(self) -> Dict:
        """당월 기준 KPI 4종 (대시보드 _refresh_kpis와 동일 로직)."""
        current_month = datetime.now().date().replace(day=1).isoformat()
        monthly = self.data_manager.get_monthly_production_stats(months=1)
        records, amount = 0, 0.0
        for r in monthly:
            if r.get("year_month") == current_month[:7]:
                records = int(r.get("record_count") or 0)
                amount = float(r.get("total_amount") or 0)
                break
        workers = self.data_manager.get_worker_stats(start_date=current_month)
        recipes = self.data_manager.get_recipe_frequency(limit=10000)
        return {"records": records, "amount": amount,
                "workers": len(workers), "recipes": len(recipes)}

    def _excel_to_pdf(self, excel_path: str, pdf_path: str) -> None:
        """win32com으로 엑셀을 일반 PDF로 변환 (스캔효과 없음)."""
        import win32com.client

        excel = None
        workbook = None
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(os.path.abspath(excel_path))
            workbook.Worksheets(1).ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        finally:
            try:
                if workbook:
                    workbook.Close(SaveChanges=False)
            except (OSError, IOError):
                logger.warning("Excel workbook 닫기 실패")
            try:
                if excel:
                    excel.Quit()
            except (OSError, IOError):
                logger.warning("Excel 프로세스 종료 실패")

    @staticmethod
    def _period_label(start_date: Optional[str], end_date: Optional[str]) -> str:
        if not start_date and not end_date:
            return "전체"
        return f"{start_date}~{end_date}"
