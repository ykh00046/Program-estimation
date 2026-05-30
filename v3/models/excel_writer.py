"""Excel 작성/서식 책임 (PDCA #22).

배합 기록을 템플릿 기반 Excel 파일로 출력한다. openpyxl만 의존하며
PDF/win32com/fitz와 분리되어 단독 테스트가 가능하다.
"""
import os
import shutil
import warnings
from typing import Any, Dict, Optional

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Border, Side, Alignment

from utils.logger import logger


class ExcelWriter:
    """엑셀 템플릿 채우기/서식 전담."""

    def __init__(self, excel_folder: str, template_file: str, cell_mapping: Dict) -> None:
        self.excel_folder = excel_folder
        self.template_file = template_file
        self.cell_mapping = cell_mapping

    def export_to_excel(self, data: Dict, include_image: bool = False,
                        image_path: Optional[str] = None,
                        include_work_time: bool = True) -> Optional[str]:
        """배합 기록을 엑셀 파일로 출력"""
        try:
            if not os.path.exists(self.template_file):
                logger.warning(f"템플릿 파일이 없습니다: {self.template_file}")
                return None

            output_file = os.path.join(self.excel_folder, f"{data['product_lot']}.xlsx")
            shutil.copy2(self.template_file, output_file)

            with warnings.catch_warnings():
                warnings.filterwarnings(
                    "ignore",
                    message="wmf image format is not supported*",
                    category=UserWarning,
                    module="openpyxl",
                )
                wb = load_workbook(output_file)
            ws = wb.active

            # 데이터 입력
            self._fill_excel_data(ws, data, include_work_time)

            if include_image and image_path:
                if not self._add_image_to_worksheet(ws, image_path):
                    logger.warning("이미지 삽입에 실패했지만 엑셀 파일은 정상 생성됩니다.")

            data_end_row = self.cell_mapping.get('data_start_row', 7) + len(data.get('materials', [])) - 1
            self._format_worksheet(ws, data_end_row)

            wb.save(output_file)
            wb.close()
            logger.info(f"엑셀 파일 생성 완료: {output_file}")
            return output_file
        except (OSError, ValueError, KeyError) as e:
            logger.error(f"엑셀 출력 오류: {e}", exc_info=True)
            return None

    def _fill_excel_data(self, ws: Any, data: Dict, include_work_time: bool = True) -> None:
        """엑셀 시트에 데이터 채우기"""
        if 'date' in self.cell_mapping: ws[self.cell_mapping['date']] = f"작업일: {data.get('work_date', '')}"
        if 'scale' in self.cell_mapping: ws[self.cell_mapping['scale']] = f"저울: {data.get('scale', '')}"
        if 'worker' in self.cell_mapping: ws[self.cell_mapping['worker']] = f"작업자 : {data.get('worker', '')}"

        if include_work_time:
            if 'work_time' in self.cell_mapping: ws[self.cell_mapping['work_time']] = f"작업시간 : {data.get('work_time', '')}"
        else:
            if 'work_time' in self.cell_mapping: ws[self.cell_mapping['work_time']] = ""

        if 'product_lot' in self.cell_mapping: ws[self.cell_mapping['product_lot']] = data.get('product_lot', '')
        if 'total_amount' in self.cell_mapping: ws[self.cell_mapping['total_amount']] = data.get('total_amount', 0) / 100

        start_row = self.cell_mapping.get('data_start_row', 7)
        for idx, material in enumerate(data.get('materials', [])):
            row = start_row + idx
            if 'material_name_col' in self.cell_mapping: ws[f"{self.cell_mapping['material_name_col']}{row}"] = material.get('material_name', '')
            if 'material_lot_col' in self.cell_mapping: ws[f"{self.cell_mapping['material_lot_col']}{row}"] = material.get('material_lot', '')
            if 'ratio_col' in self.cell_mapping: ws[f"{self.cell_mapping['ratio_col']}{row}"] = material.get('ratio', 0)
            if 'theory_amount_col' in self.cell_mapping: ws[f"{self.cell_mapping['theory_amount_col']}{row}"] = material.get('theory_amount', 0)
            if 'actual_amount_col' in self.cell_mapping: ws[f"{self.cell_mapping['actual_amount_col']}{row}"] = material.get('actual_amount', 0)

    def _format_worksheet(self, ws: Any, data_end_row: int) -> None:
        """워크시트 서식 정리"""
        self._delete_empty_rows(ws, data_end_row)
        self._apply_cell_merges(ws, data_end_row)
        self._apply_borders(ws, data_end_row)

    def _add_image_to_worksheet(self, ws: Any, image_path: str) -> bool:
        """워크시트 G2 셀에 이미지 추가"""
        try:
            if not os.path.exists(image_path):
                logger.warning(f"이미지 파일이 존재하지 않습니다: {image_path}")
                return False
            img = OpenpyxlImage(image_path)
            img.width = 228
            img.height = 65
            img.anchor = "G2"
            ws.add_image(img)
            return True
        except (OSError, IOError) as e:
            logger.error(f"이미지 추가 중 오류: {e}", exc_info=True)
            return False

    def _delete_empty_rows(self, ws: Any, data_end_row: int) -> None:
        """불필요한 빈 행 삭제"""
        try:
            for row_num in range(ws.max_row, data_end_row, -1):
                if all(ws.cell(row=row_num, column=col).value is None for col in range(1, 8)):
                    ws.delete_rows(row_num, 1)
        except (OSError, IOError) as e:
            logger.warning(f"행 삭제 중 오류: {e}")

    def _apply_cell_merges(self, ws: Any, data_end_row: int) -> None:
        """셀 병합 적용"""
        try:
            ws.merge_cells(f'A6:A{data_end_row}')
            ws.merge_cells(f'B6:B{data_end_row}')
            ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
            ws['B6'].alignment = Alignment(horizontal='center', vertical='center')
        except (OSError, ValueError, RuntimeError) as e:
            logger.warning(f"셀 병합 중 오류: {e}")

    def _apply_borders(self, ws: Any, data_end_row: int) -> None:
        """테이블 경계선 적용"""
        try:
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in range(5, data_end_row + 1):
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = thin_border
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        except (OSError, ValueError, RuntimeError) as e:
            logger.warning(f"경계선 적용 중 오류: {e}")
